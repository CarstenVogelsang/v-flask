"""XLSX Export Service for Fragebogen responses.

Exports all participation data to Excel format with:
- Participant information (name, email, type)
- Timestamps (invited, started, completed)
- All questions as columns
- Answers formatted by question type

Requires: openpyxl (install with: pip install v-flask[export])

Example:
    from v_flask_plugins.fragebogen.services import get_export_service

    service = get_export_service()
    xlsx_buffer = service.export_to_xlsx(fragebogen)

    # In Flask route:
    return Response(
        xlsx_buffer.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="export.xlsx"'}
    )
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from typing import TYPE_CHECKING, Any, Generator

from v_flask.extensions import db
from v_flask_plugins.fragebogen.models import (
    Fragebogen,
    FragebogenTeilnahme,
    TeilnahmeStatus,
)

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ExportOptions:
    """Options for XLSX export.

    Attributes:
        include_timestamps: Include invitation/start/completion timestamps.
        include_incomplete: Include participations that aren't completed yet.
        date_format: Strftime format for date/time columns.
    """
    include_timestamps: bool = True
    include_incomplete: bool = False
    date_format: str = '%d.%m.%Y %H:%M'


class FragebogenExportService:
    """Service for exporting Fragebogen responses to XLSX.

    Memory-efficient implementation using generator patterns.
    Supports all question types with appropriate formatting.
    """

    def export_to_xlsx(
        self,
        fragebogen: Fragebogen,
        options: ExportOptions | None = None
    ) -> BytesIO:
        """Export all responses to XLSX format.

        Args:
            fragebogen: The Fragebogen to export.
            options: Export configuration options.

        Returns:
            BytesIO buffer containing the XLSX file.

        Raises:
            ImportError: If openpyxl is not installed.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl ist nicht installiert. "
                "Installiere mit: pip install v-flask[export]"
            )

        options = options or ExportOptions()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Antworten'

        # Build header row
        headers = self._build_headers(fragebogen, options)

        # Style header row
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color='E0E0E0',
            end_color='E0E0E0',
            fill_type='solid'
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Write data rows
        row_idx = 2
        for teilnahme in self._get_teilnahmen(fragebogen, options):
            row_data = self._build_row(teilnahme, fragebogen, options)
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1

        # Auto-adjust column widths (with max limit)
        self._adjust_column_widths(ws, headers)

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _build_headers(
        self,
        fragebogen: Fragebogen,
        options: ExportOptions
    ) -> list[str]:
        """Build header row with participant info and question columns."""
        headers = [
            'ID',
            'Name',
            'E-Mail',
            'Typ',
            'Status',
        ]

        if options.include_timestamps:
            headers.extend([
                'Eingeladen am',
                'Gestartet am',
                'Abgeschlossen am',
            ])

        # Add question columns
        for frage in fragebogen.fragen:
            frage_typ = frage.get('typ', 'text')

            # Skip group questions (they are containers)
            if frage_typ == 'group':
                # Add sub-fields as columns
                for sub_field in frage.get('fields', []):
                    field_label = sub_field.get('label', sub_field.get('id', ''))
                    headers.append(f"{frage.get('frage', '')}: {field_label}")
            else:
                headers.append(frage.get('frage', frage.get('id', '')))

        return headers

    def _get_teilnahmen(
        self,
        fragebogen: Fragebogen,
        options: ExportOptions
    ) -> Generator[FragebogenTeilnahme, None, None]:
        """Get participations to export (generator for memory efficiency)."""
        query = db.session.query(FragebogenTeilnahme).filter_by(
            fragebogen_id=fragebogen.id
        )

        if not options.include_incomplete:
            query = query.filter_by(status=TeilnahmeStatus.ABGESCHLOSSEN.value)

        query = query.order_by(FragebogenTeilnahme.id)

        # Use yield_per for memory efficiency with large datasets
        for teilnahme in query.yield_per(100):
            yield teilnahme

    def _build_row(
        self,
        teilnahme: FragebogenTeilnahme,
        fragebogen: Fragebogen,
        options: ExportOptions
    ) -> list[Any]:
        """Build data row for a single participation."""
        row = [
            teilnahme.id,
            teilnahme.display_name,
            teilnahme.kontakt_email or '',
            teilnahme.teilnehmer_typ or 'anonym',
            self._translate_status(teilnahme.status),
        ]

        if options.include_timestamps:
            date_fmt = options.date_format
            row.extend([
                teilnahme.einladung_gesendet_am.strftime(date_fmt) if teilnahme.einladung_gesendet_am else '',
                teilnahme.gestartet_am.strftime(date_fmt) if teilnahme.gestartet_am else '',
                teilnahme.abgeschlossen_am.strftime(date_fmt) if teilnahme.abgeschlossen_am else '',
            ])

        # Add answers
        for frage in fragebogen.fragen:
            frage_typ = frage.get('typ', 'text')

            if frage_typ == 'group':
                # Handle group fields
                for sub_field in frage.get('fields', []):
                    field_id = f"{frage['id']}.{sub_field.get('id', '')}"
                    antwort = teilnahme.get_antwort(field_id)
                    row.append(self._format_answer(antwort, sub_field.get('typ', 'text')))
            else:
                antwort = teilnahme.get_antwort(frage['id'])
                row.append(self._format_answer(antwort, frage_typ))

        return row

    def _format_answer(self, antwort, frage_typ: str) -> str:
        """Format answer value based on question type."""
        if not antwort:
            return ''

        value = antwort.value
        if value is None:
            return ''

        # Type-specific formatting
        if frage_typ == 'multiple_choice':
            if isinstance(value, list):
                return ', '.join(str(v) for v in value)
            return str(value)

        elif frage_typ == 'ja_nein':
            if value is True:
                return 'Ja'
            elif value is False:
                return 'Nein'
            return str(value)

        elif frage_typ == 'date':
            # Try to format as German date
            try:
                if isinstance(value, str):
                    dt = datetime.fromisoformat(value)
                    return dt.strftime('%d.%m.%Y')
            except (ValueError, TypeError):
                pass
            return str(value)

        elif frage_typ == 'dropdown':
            # Check for freitext
            freitext = antwort.freitext
            if freitext and value in ('Sonstiges', 'Andere', 'Other'):
                return freitext
            return str(value)

        elif frage_typ == 'table':
            # Export table as JSON string
            if value:
                return json.dumps(value, ensure_ascii=False)
            return ''

        elif frage_typ == 'skala':
            # Return numeric value as is
            return str(value) if value is not None else ''

        # Default: convert to string
        return str(value) if value else ''

    def _translate_status(self, status: str) -> str:
        """Translate status to German."""
        translations = {
            'eingeladen': 'Eingeladen',
            'gestartet': 'Gestartet',
            'abgeschlossen': 'Abgeschlossen',
        }
        return translations.get(status, status)

    def _adjust_column_widths(self, ws: 'Worksheet', headers: list[str]) -> None:
        """Auto-adjust column widths with reasonable limits."""
        from openpyxl.utils import get_column_letter

        for col_idx, header in enumerate(headers, start=1):
            # Calculate width based on header length (with min/max limits)
            width = min(max(len(header) + 2, 12), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = width


# =============================================================================
# Singleton
# =============================================================================

_export_service: FragebogenExportService | None = None


def get_export_service() -> FragebogenExportService:
    """Get the export service singleton.

    Returns:
        The shared FragebogenExportService instance.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    global _export_service
    if _export_service is None:
        _export_service = FragebogenExportService()
    return _export_service


# =============================================================================
# Exports
# =============================================================================

__all__ = [
    'FragebogenExportService',
    'get_export_service',
    'ExportOptions',
]
